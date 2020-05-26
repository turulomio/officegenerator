## @namespace officegenerator.libxlsxgenerator
## @brief Este módulo permite la lectura y escritura de ficheros xlsx de Microsoft Excel
##
## You can change current sheet, with createSheet or using setCurrentSheet. After that all commands will use that sheet until you change it again
import gettext
import openpyxl
import openpyxl.comments
import openpyxl.cell
import openpyxl.styles
import openpyxl.worksheet
import openpyxl.worksheet.worksheet
import openpyxl.formatting.rule
import pkg_resources

from decimal import Decimal
from officegenerator.commons import columnAdd, Coord, Range, topLeftCellNone, convert_command, generate_formula_total_string, Coord_from_index
from officegenerator.objects.currency import Currency, currency_symbol
from officegenerator.objects.formula import isFormula
from officegenerator.objects.percentage import Percentage
from os import path, makedirs
from tempfile import TemporaryDirectory

try:
    t=gettext.translation('officegenerator',pkg_resources.resource_filename("officegenerator","locale"))
    _=t.gettext
except:
    _=str

class ColumnWidthXLSX:
    Date=40
    Detetime=60

class XLSX_Commons:
    def __init__(self):
        pass

    ## It returns a sheet object with the index id
    def get_sheet_by_id(self, id):
        return self.wb[self.wb.sheetnames[id]]

    ## It returns a index integer of the sheet with a given name
    def get_sheet_id(self, name):
        for id, s_name in enumerate(self.wb.sheetnames):
            if s_name==name:
                return id
        return None

    ## Function that establishes current worksheet. Updates self.ws_current and self.ws_current_id
    ##
    ## id Is a integer beginning with 0
    ## name is the title of the sheet
    ## @param id_or_name Index or Nmae
    def setCurrentSheet(self, id_or_name):
        if id_or_name.__class__==int:
            self.ws_current_id=id_or_name
        else:#name
            self.ws_current_id=self.get_sheet_id(id_or_name)
        self.ws_current=self.get_sheet_by_id(self.ws_current_id)

    ## @param sheet_index Integer index of the sheet
    ## @param range_ Range object to get OdfCell. If None returns all OdfCell from sheet
    ## @return Returns a list of rows of object values
    def cells(self, sheet_index, range_=None):
        if range_ is None:
            range_=self.getSheetRange(sheet_index)
        else:
            range_=Range.assertRange(range_)
        r=[]
        for row in range(range_.numRows()):
            tmprow=[]
            for column in range(range_.numColumns()):
                tmprow.append(self.getCell(sheet_index, range_.start.addRowCopy(row).addColumnCopy(column)))
            r.append(tmprow)
        return r
        
    ## Returns a list of rows with the values of the sheet
    ## @param sheet_index Integer index of the sheet
    ## @param skip_up int. Number of rows to skip at the begining of the list of rows (lor)
    ## @param skip_down int. Number of rows to skip at the end of the list of rows (lor)
    ## @return Returns a list of rows of object values
    def values(self, sheet_index, skip_up=0, skip_down=0):
        r=[]
        for row in range(skip_up, self.rowNumber(sheet_index)-skip_down):
            tmprow=[]
            for column in range(self.columnNumber(sheet_index)):
                tmprow.append(self.getCellValue(sheet_index, Coord("A1").addRow(row).addColumn(column)))
            r.append(tmprow)
        return r
 
    ## @param sheet_index Integer index of the sheet
    ## @param range_ Range object to get values. If None returns all values from sheet
    ## @return Returns a list of rows of object values
    def values_by_range(self, sheet_index, range_):
        range_=Range.assertRange(range_)
        r=[]
        for row in range(range_.numRows()):
            tmprow=[]
            for column in range(range_.numColumns()):
                tmprow.append(self.getCellValue(sheet_index, range_.start.addRowCopy(row).addColumnCopy(column)))
            r.append(tmprow)
        return r
    
    ## @param sheet_index Integer index of the sheet
    ## @param column_letter Letter of the column to get values
    ## @param skip Integer Number of top rows to skip in the result
    ## @return List of values
    def getColumnValues(self, sheet_index, column_letter, skip_up=0, skip_down=0):
        r=[]
        for row in range(skip_up, self.rowNumber(sheet_index)-skip_down):
            r.append(self.getCellValue(sheet_index, Coord(column_letter+"1").addRow(row)))
        return r    

    ## @param sheet_index Integer index of the sheet
    ## @param row_number String Number of the row to get values
    ## @param skip Integer Number of top rows to skip in the result
    ## @return List of values
    def getRowValues(self, sheet_index, row_number, skip_left=0, skip_right=0):
        r=[]
        for column in range(skip_left, self.columnNumber(sheet_index)-skip_right):
            r.append(self.getCellValue(sheet_index, Coord("A"+row_number).addColumn(column)))
        return r

    ## Return a Range object with the limits of the index sheet
    def getSheetRange(self, sheet_index):
        self.setCurrentSheet(sheet_index)
        return Range(self.ws_current.calculate_dimension())
        
    def rowNumber(self, sheet_index):
        self.setCurrentSheet(sheet_index)
        return self.ws_current.max_row
        
    def columnNumber(self, sheet_index):
        self.setCurrentSheet(sheet_index)
        return self.ws_current.max_column
        
    ## Returns the cell value
    def getCellValue(self, sheet_index, coord):
        def hasCurrency(cell):
            for currency in ["EUR", "USD","GBP"]:
                if currency in cell.number_format and cell.data_type=="n":
                    return Currency(cell.value, currency)
            return None
        #----------------------
        self.setCurrentSheet(sheet_index)
        cell=self.getCell(sheet_index, coord)
        #print(cell.number_format, cell.value, cell.data_type)
        #print(dir(cell))
        has_currency=hasCurrency(cell)
        if has_currency is not None:
            return has_currency
        elif cell.data_type=="f":#Formula
            return cell.value
        elif "%" in cell.number_format:
            return Percentage(cell.value, 1)
        elif cell.data_type=="d" and "H" not in cell.number_format.upper():#Date
            return cell.value.date()
        elif cell.data_type=="d" and "Y" not in cell.number_format.upper():#Tiime
            return cell.value
        elif cell.data_type=="d":#Datetime
            return cell.value
        elif cell.data_type == "b": #Boolean
            return cell.value
        elif cell.number_format=='"BOOL"e"AN"':
            if cell.value==1:
                return True
            elif cell.value==0:
                return False
        elif cell.data_type == "s":
            return str(cell.value)
        elif cell.value is None:
            return None
        elif cell.data_type=="n":
            try:
                return Decimal(str(cell.value))
            except:
                return str(cell.value)

    ## Returns an odfcell object
    def getCell(self, sheet_index,  coord):
        self.setCurrentSheet(sheet_index)
        coord=Coord.assertCoord(coord)
        return self.ws_current[coord.string()]

class XLSX_Write(XLSX_Commons):
    ## @param filename
    ## @param template
    ## @param data_only load_workbook now accepts data_only to allow extracting values only from formulae. Default is false
    def __init__(self,filename, template=None, data_only=False, keep_vba=True):
        XLSX_Commons.__init__(self)
        self.filename=filename
        self.template=template
        if template is None:
            self.wb=openpyxl.Workbook()
        else:
            self.wb=openpyxl.load_workbook(self.template, keep_vba=keep_vba, data_only=data_only)

        self.ws_current=self.wb.active
        self.setCurrentSheet(self.ws_current.title)

        self.stOrange=openpyxl.styles.Color('FFFFDCA8')
        self.stYellow=openpyxl.styles.Color('FFFFFFC0')
        self.stGreen=openpyxl.styles.Color('FFC0FFC0')
        self.stGrayLight=openpyxl.styles.Color('FFDCDCDC')
        self.stGrayDark=openpyxl.styles.Color('FFC3C3C3')
        self.stWhite=openpyxl.styles.Color('FFFFFFFF')
    
    ## Returns the style name of a givenven color
    ## @param openpyxl.styles.Color
    ## @return string
    def styleName(self, color):
        if color==self.stOrange:
            return "Orange"
        elif color==self.stYellow:
            return "Yellow"
        elif color==self.stGreen:
            return "Green"
        elif color==self.stGrayDark:
            return "Dark gray"
        elif color==self.stGrayLight:
            return "Light gray"
        elif color==self.stWhite:
            return "White"
        elif color==None:
            return "Normal"

    ## Freeze panels in a sheet and sets the selected cell
    ## Selects a cell https://openpyxl.readthedocs.io/en/latest/_modules/openpyxl/worksheet/views.html#Selection
    ## To DEBUG THIS function rename xlsx to zip and enter in xl/worksheet to see code and props
    ## @param freeze_coord, Cell where panels are frrozen. Can be a string or a Coord object.
    ## @param selected_coord. Cell selected opening sheet. Can be a string or a Coord object.
    ## @param topLeftCell, topleftcell to show in sheet after opening. Can be a string or a Coord object.
    def freezeAndSelect(self, freeze_coord, selected_coord=None, topleftcell_coord=None):
        if selected_coord is None:
            selected_coord=Coord_from_index(self.columnNumber(self.ws_current_id)-1, self.rowNumber(self.ws_current_id)-1)

        if topleftcell_coord is None:
            topleftcell_coord=topLeftCellNone(freeze_coord, selected_coord)

        freeze_coord=Coord.assertCoord(freeze_coord)
        selected_coord=Coord.assertCoord(selected_coord)
        topleftcell_coord=Coord.assertCoord(topleftcell_coord)
        self.ws_current.freeze_panes=self.ws_current[freeze_coord.string()]
        sheet=self.ws_current.views.sheetView[0]
        if freeze_coord.letterIndex()>0 and freeze_coord.numberIndex()>0:#Freeze C3 WORKS
            sheet.selection[2].activeCell=selected_coord.string()
            sheet.selection[2].sqref=selected_coord.string()
            sheet.pane.activePane = 'bottomRight'
            sheet.pane.topLeftCell=topleftcell_coord.string()
        elif freeze_coord.letterIndex()==0 and freeze_coord.numberIndex()>0:#Freeze A3  WORKS
            topLeftCellFirst=topleftcell_coord.letter+"1"
            sheet.topLeftCell=topLeftCellFirst
            sheet.view="normal"
            sheet.pane.xSplit="0"
            sheet.selection[0].pane="bottomLeft"
            sheet.selection[0].activeCell=selected_coord.string()
            sheet.selection[0].sqref=selected_coord.string()
            sheet.selection[0].state="frozen"
            sel = list(sheet.selection)
            sel.insert(0, openpyxl.worksheet.worksheet.Selection(pane="topLeft", activeCell=topLeftCellFirst, sqref=topLeftCellFirst))
            sheet.selection = sel
            sheet.pane.topLeftCell=topleftcell_coord.string()
        elif freeze_coord.letterIndex()>0 and freeze_coord.numberIndex()==0:#Freeze C1 WORKS
            #Comparing output with a officegenerator good xlsx
            topLeftCellFirst="A"+topleftcell_coord.number
            sheet.topLeftCell=topLeftCellFirst
            sheet.pane.ySplit="0"
            sheet.selection[0].activeCell=selected_coord.string()
            sheet.selection[0].sqref=selected_coord.string()
            sel = list(sheet.selection)
            sel.insert(0, openpyxl.worksheet.worksheet.Selection(pane="topLeft", activeCell=topLeftCellFirst, sqref=topLeftCellFirst))
            sheet.selection = sel
            sheet.pane.topLeftCell=topleftcell_coord.string()
        elif freeze_coord.letterIndex()==0 and freeze_coord.numberIndex()==0:#Freeze A1 WORKS
            sheet.selection[0].activeCell=selected_coord.string()
            sheet.selection[0].sqref=selected_coord.string()
            sheet.selection[0].pane='topLeft'
            sheet.topLeftCell=topleftcell_coord.string()

    ## Changes name of the current sheet
    def setSheetName(self, name):
        self.ws_current.title=name

    ## Create a sheet at the end, renames it and selects it as current
    def createSheet(self, name):
        self.wb.create_sheet(title=name)
        self.setCurrentSheet(name)

    def setColorScale(self, range):
        self.ws_current.conditional_formatting.add(range, 
                            openpyxl.formatting.rule.ColorScaleRule(
                                                start_type='percentile', start_value=0, start_color='00FF00',
                                                mid_type='percentile', mid_value=50, mid_color='FFFFFF',
                                                end_type='percentile', end_value=100, end_color='FF0000'
                                                )
                                            )
    ## Returns sheet_name
    def sheet_name(self, id=None):
        if id==None:
            id=self.ws_current_id
        return self.wb.sheetnames[id]

    ## After removing it sets current sheet to 0 index
    def remove_sheet_by_id(self, id):
        ws=self.get_sheet_by_id(id)
        self.wb.remove(ws)
        self.setCurrentSheet(0)

    def save(self, filename=None):
        if filename==None:
            filename=self.filename
        if path.dirname(filename)!="":
            makedirs(path.dirname(filename), exist_ok=True)
        self.wb.save(filename)

        if path.exists(filename)==False:
            print(_("*** ERROR: File wasn't generated ***"))

    ## Internal function to set the number format
    ##
    ## This strings are openpyxl string not libreoffice cell string
    ## @param cell is a cell object
    ## @param value Value to add to the cell
    ## @param style Color or None. If None this function it's ignored
    ## @param decimals Number of decimals
    def __setNumberFormat(self, cell, value, style, decimals):     
        if style==None:
            return
        if value.__class__.__name__ in ("int", ):#Un solo valor
            cell.number_format='#,##0;[RED]-#,##0'
        elif value.__class__.__name__ in ("float", "Decimal"):#Un solo valor
            zeros=decimals*"0"
            cell.number_format="#,##0.{0};[RED]-#,##0.{0}".format(zeros)
        elif value.__class__.__name__ in ("datetime", ):
            cell.number_format="YYYY-MM-DD HH:mm"
        elif value.__class__.__name__ in ("date", ):
            cell.number_format="YYYY-MM-DD"
        elif value.__class__.__name__ in ("Currency", "Money" ):
            cell.number_format='#,##0.00 "{0}";[RED]-#,##0.00 "{0}"'.format(currency_symbol(value.currency))
        elif value.__class__.__name__ in ("Percentage",  ):
            cell.number_format="#.##0,00 %;[RED]-#.##0,00 %"
        elif value.__class__.__name__ =="bool":
            cell.number_format="BOOLEAN"

    ## Internal function to set the number format of a formula
    ##
    ## This strings are openpyxl string not libreoffice cell string
    ## @param cell is a cell object
    ## @param resultclass int, float, Decimal, datetime.datetime, datetime.date,"€","$" (Currency.symbol), Percentage
    ##    Currency it's not used because value is a string, and I can't get value.currency. Maybe I should create a Formula object, but Iwill have to change a lot of apps
    ## @param style Color or None. If None this function it's ignored
    ## @param decimals Number of decimals
    def __setFormulaNumberFormat(self, cell, value, resultclass, style, decimals):     
        if style==None:
            return
        if resultclass in ("int", ):#Un solo valor
            cell.number_format='#,##0;[RED]-#,##0'
        elif resultclass in ("float", "Decimal"):#Un solo valor
            zeros=decimals*"0"
            cell.number_format="#,##0.{0};[RED]-#,##0.{0}".format(zeros)
        elif resultclass in ("datetime", ):
            cell.number_format="YYYY-MM-DD HH:mm"
        elif resultclass in ("date", ):
            cell.number_format="YYYY-MM-DD"
        elif resultclass.__class__ in ("str", ):
            cell.number_format='#,##0.00 "{0}";[RED]-#,##0.00 "{0}"'.format(resultclass)
        elif resultclass in ("Percentage", ):
            cell.number_format="#.##0,00 %;[RED]-#.##0,00 %"
        elif resultclass in ("$", "€"):
            cell.number_format='#,##0.00 "{0}";[RED]-#,##0.00 "{0}"'.format(resultclass)
        elif resultclass=="bool":
            cell.number_format="BOOLEAN"

    ## Internat function to set a cell. All properties except border that it's setted in overwrite functions (merged and no merged)
    ## @param cell is a cell object
    def __setValue(self, cell, value, style, decimals, alignment):     
        if value==None:
            return
        elif value.__class__.__name__ in ("Currency", "Money" ):
            cell.value=value.amount
        elif value.__class__.__name__ in ("Percentage", ):
            cell.value=value.value
        elif value.__class__.__name__ == "bool":
            cell.value=value
        else:
            cell.value=value

    ## Internal method to set a not merged cell
    ## @param cell is a cell object
    ## @param value Value to add to the cell
    ## @param style Color or None. If None this function it's ignored
    ## @param decimals Number of decimals
    ## @param alignment Cell alignment
    def __setCell(self, coord, value, style=None, decimals=2, alignment=None):
        coord=Coord.assertCoord(coord)
        cell=self.getCell(self.ws_current_id, coord.string())
        self.__setValue(cell, value, style, decimals, alignment)
        self.__setBorder(cell, style)
        self.__setAlignment(cell, value, style, alignment)
        self.__setNumberFormat(cell, value, style, decimals)      

        if style!=None:
            cell.fill=openpyxl.styles.PatternFill("solid", fgColor=style)
            bold=False if style==self.stWhite else True
            cell.font=openpyxl.styles.Font(name='Arial', size=10, bold=bold)

    ## Internat function to set cell alignment
    ## @param cell is a cell object
    ## @style Color or None. This method is ignored if style=None
    def __setAlignment(self, cell, value,  style, alignment):  
        if style==None:
            return
        if alignment==None:
            if value.__class__.__name__ in ("str", ):#Un solo valor
                if value.startswith("=") or value.startswith("+"):
                    alignment='right'
                else:
                    alignment='left'
            else:
                alignment='right'
        cell.alignment=openpyxl.styles.Alignment(horizontal=alignment, vertical='center')

    ## Writes a cell or a list of cell or a list of list of cells
    ## @param coord Can be a Coord or a string with text coords
    ## @param result Can be a value, a list of values or a list of lists of values
    ## @param style its a openpyxl.styles.Color object. There are several predefined stGreen, stGrayDark, stGrayLight, stOrange, stYellow, stWhite or None. None is used to preserve template cell and the value is the only thing will be changed
    ## @param decimals Integer with the number of decimals. 2 by default
    ## @param alignment String None by default. Can be "right","left","center"
    def overwrite(self, coord, result, style=None,  decimals=2, alignment=None):
        coord=Coord.assertCoord(coord)
        if result.__class__== list:#Una lista
            for i,row in enumerate(result):
                if row.__class__ in (list, ):#Una lista de varias columnas
                    for j,column in enumerate(row):
                        self.__setCell(Coord(coord.string()).addRow(i).addColumn(j), result[i][j], style, decimals, alignment )   
                else:#Una lista de una columna
                    self.__setCell(Coord(coord.string()).addRow(i), result[i], style, decimals, alignment )
        else:#Un solo valor
            self.__setCell(coord, result, style, decimals, alignment )
            

    ## Writes a formula in a cell 
    ## @param coord Can be a Coord or a string with text coord
    ## @param result Can be a value
    ## @param resultclass string that can be: int, float, Decimal, datetime.datetime, datetime.date, Currency, Percentage
    ## @param style its a openpyxl.styles.Color object. There are several predefined stGreen, stGrayDark, stGrayLight, stOrange, stYellow, stWhite or None. None is used to preserve template cell and the value is the only thing will be changed
    ## @param decimals Integer with the number of decimals. 2 by default
    ## @param alignment String None by default. Can be "right","left","center"
    def overwrite_formula(self, coord, value, resultclass=None, style=None, decimals=2, alignment=None):
        if isFormula(value)==False:
            print(_("This is not a formula. You can't use overwrite_formula"))
            return
        if value.__class__==list:
            print("Adding formula list is not allowed")
            return
        coord=Coord.assertCoord(coord)
        cell=self.getCell(self.ws_current_id, coord)       
        
        self.__setValue(cell, value, style, decimals, alignment)
        self.__setBorder(cell, style)
        self.__setAlignment(cell, value, style, alignment)
        self.__setFormulaNumberFormat(cell, value, resultclass, style, decimals)      

        if style!=None:
            cell.fill=openpyxl.styles.PatternFill("solid", fgColor=style)
            bold=False if style==self.stWhite else True
            cell.font=openpyxl.styles.Font(name='Arial', size=10, bold=bold)
            
            
    ## @param cood Coord from we are going to add totals
    ## @param list_of_totals List with strings or keys. Example: ["Total", "#SUM", "#AVG"]...
    ## @param list_of_styles List with string styles or None. If none tries to guest from top column object. List example: ["GrayLightPercentage", "GrayLightInteger"]
    ## @param string with the row where th3e total begins
    ## @param string with the rew where the formula ends. If None it's a coord.row -1
    def overwriteTotalsHorizontal(self, coord, list_of_totals, list_of_styles=None, row_from="2", row_to=None,  list_of_decimals=2):
        coord=Coord.assertCoord(coord)
        for letter, total in enumerate(list_of_totals):
            coord_total=coord.addColumnCopy(letter)
            coord_total_from=Coord(coord_total.letter+row_from)
            if row_to is None:
                coord_total_to=coord_total.addRowCopy(-1)# row above
            else:
                coord_total=Coord(coord_total.letter+row_to)

            if list_of_styles is None:
                style=self.stGrayLight
            else:
                style=list_of_styles[letter]
                
            if list_of_decimals.__class__.__name__=="int":
                decimals=2
            else:
                decimals=list_of_decimals[letter]

            self.overwrite(coord_total, generate_formula_total_string(total, coord_total_from, coord_total_to),  style, decimals)
            
    ## @param cood Coord from we are going to add totals
    ## @param list_of_totals List with strings or keys. Example: ["Total", "#SUM", "#AVG"]...
    ## @param list_of_styles List with string styles or None. If none tries to guest from top column object. List example: ["GrayLightPercentage", "GrayLightInteger"]
    ## @param string with the row where th3e total begins
    ## @param string with the rew where the formula ends. If None it's a coord.row -1
    def overwriteTotalsVertical(self, coord, list_of_totals, list_of_styles=None, column_from="B", column_to=None,  list_of_decimals=2):
        coord=Coord.assertCoord(coord)
        for i, total in enumerate(list_of_totals):
            coord_total=coord.addRowCopy(i)
            coord_total_from=Coord(column_from+coord_total.number)
            if column_to is None:
                coord_total_to=coord_total.addColumnCopy(-1)# row above
            else:
                coord_total=Coord(column_to+coord_total.number)

            if list_of_styles is None:
                style=self.stGrayLight
            else:
                style=list_of_styles[i]

            if list_of_decimals.__class__.__name__=="int":
                decimals=2
            else:
                decimals=list_of_decimals[i]


            self.overwrite(coord_total, generate_formula_total_string(total, coord_total_from, coord_total_to),  style, decimals)


    ##Sets border to a cell not merged
    ## @param cell is a cell object
    ## @param style Color or None. If None this function it's ignored
    def __setBorder(self, cell, style):
        if style==None:
            return
        cell.border=openpyxl.styles.Border(
            left=openpyxl.styles.Side(border_style='thin'),
            top=openpyxl.styles.Side(border_style='thin'),
            right=openpyxl.styles.Side(border_style='thin'),
            bottom=openpyxl.styles.Side(border_style='thin') 
        )

    ## Sets cell name to use in formulas. Fails if range_string is not fixed. For example: $A$4
    ## @param range_string 
    ## @param name
    def setCellName(self, range_string, name):
        self.wb.create_named_range(name, self.ws_current, range_string)

    ## Set columns width in current sheet
    ## @param arrWidths List with integers representing column width
    def setColumnsWidth(self, arrWidths):
        for i in range(len(arrWidths)):
            self.ws_current.column_dimensions[columnAdd("A", i)].width=arrWidths[i]

    ## Create a merged cell
    ## @param range_string Can be a Range or a range string 
    ## @param value Can be a value. Must be the value only for the first cell
    ## @param style its a openpyxl.styles.Color object. There are several predefined stGreen, stGrayDark, stGrayLight, stOrange, stYellow, stWhite or None. None is used to preserve template cell and the value is the only thing will be changed
    ## @param decimals Integer with the number of decimals. 2 by default
    ## @param alignment String None by default. Can be "right","left","center"
    def overwrite_and_merge(self, range,  value, style=None,  decimals=2, alignment=None):
        range=Range.assertRange(range)
        self.ws_current.merge_cells(range.string())
        top = openpyxl.styles.Border(top=openpyxl.styles.Side(border_style='thin'))
        left = openpyxl.styles.Border(left=openpyxl.styles.Side(border_style='thin'))
        right = openpyxl.styles.Border(right=openpyxl.styles.Side(border_style='thin'))
        bottom = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'))

        self.__setCell(range.start.string(), value, style, decimals, alignment)

        rows = self.ws_current[range.string()]

        for cell in rows[0]:
            cell.border = cell.border + top
        for cell in rows[-1]:
            cell.border = cell.border + bottom

        for row in rows:
            l = row[0]
            r = row[-1]
            l.border = l.border + left
            r.border = r.border + right

    ## Sets a comment
    ## @param strcell String "A1" for example
    def setComment(self, coord_string, comment):
        self.ws_current[coord_string].comment=openpyxl.comments.Comment(comment, "PySGAE")

class XLSX_Read(XLSX_Commons):
    ## @param filename
    ## @param data_only load_workbook now accepts data_only to allow extracting values only from formulae. Default is False to avoid lose information
    ## @param keep_vba. Default is True
    def __init__(self, filename, data_only=False, keep_vba=True):
        XLSX_Commons.__init__(self)
        self.filename=filename
        self.wb=openpyxl.load_workbook(self.filename, keep_vba=keep_vba, data_only=data_only)
        self.ws_current=self.wb.active
        self.setCurrentSheet(self.ws_current.title)


## Gets a XLSX file and rewrites it with libreoffice convert-to command
## Can be used to assign data values formulas to file. Or to fix ploblems on specific files.
## I had problems with concurrency due to convrert-to opens a file and locks it. I should not use convert-to with concurrency
## @param filename_from canbe an ods or a xlsx file.
## @param filename_to Must be a xlsx file path
## Returns the name of the recently created file
def create_rewritten_xlsx(filename_from, filename_to=None):
    if filename_to is None:
        filename_to=filename_from + ".rewritten.xlsx"
    convert_command(filename_from, filename_to)
    return filename_to


## @param filename_from canbe an ods or a xlsx file.
## @param filename_to Must be a xlsx file path
## Returns the name of the recently created file
def create_data_only_xlsx(filename_from, filename_to=None):
    if filename_to is None:
        filename_to=filename_from + ".data_only.xlsx"
    with TemporaryDirectory(prefix="officegenerator_") as tmp_name:
        temporal_path="{}/{}.rewritten.xlsx".format(tmp_name, path.basename(filename_to))
        convert_command(filename_from, temporal_path)
        xlsx=XLSX_Write(filename_to, template=temporal_path, data_only=True)
        xlsx.save()
    return filename_to
        
